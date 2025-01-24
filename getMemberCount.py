'''
This script retrieves all members and outside collaborators from the organizations listed in the .env file 
by interacting with the GitHub API. It gathers member details, processes the data, and outputs the results 
into an Excel file with the current date in its filename.

Main steps:
1. Read organization names and tokens from the .env file.
2. For each organization, retrieve:
   - All members using the GitHub Members API.
   - All outside collaborators using the GitHub Outside Collaborators API.
3. Merge the retrieved data and compare it with previous records to update missing display names.
4. Convert organization names from English to Chinese using a reference table.
5. Identify each member's department based on predefined rules and mappings.
6. Generate summary reports with member counts and department-wise statistics.
7. Save the final processed data to an Excel file with proper formatting and logging.
8. Log the entire execution process, including errors, for easier monitoring and debugging.

'''

import requests
import time
import logging
import pandas as pd
import re
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

DIR = os.getcwd()
CURRENTDAY = datetime.now().strftime("%Y%m%d")

MEMBER_FOLDER = os.path.abspath(os.path.join(DIR, f"history_file/members"))
MEMBER_FILE = os.path.abspath(os.path.join(MEMBER_FOLDER, f"各組織members_{CURRENTDAY}.xlsx"))
MEMBER_COUNT_FILE = os.path.abspath(os.path.join(MEMBER_FOLDER, f'各部門member_count_{CURRENTDAY}.xlsx'))

LOG_FOLDER = os.path.abspath(os.path.join(DIR, f"history_file/members/log"))
LOG_FILE = os.path.abspath(os.path.join(LOG_FOLDER, f"members_{CURRENTDAY}.log"))

xlsx_files = [int(f[11:19]) for f in os.listdir(MEMBER_FOLDER) if f.startswith('各組織')]
LATEST = max(xlsx_files)

N = 100
N_DAYS_AGO = time.time() - (N * 86400)


logging.basicConfig(filename=LOG_FILE, level=logging.INFO,encoding='utf-8',
                    format='%(asctime)s - %(levelname)s - %(message)s')

df_dept = pd.read_excel(os.path.join(DIR, '參考檔案\組織中英對照表.xlsx'))
df_before = pd.read_excel(f'history_file\members\各組織members_{LATEST}.xlsx')


def fetch_all_members(orgName, token):
    url = f"https://api.github.com/orgs/{orgName}/members"
    headers = {
        'Authorization': f'token {token}'
    }
    members = []
    page = 1  # 分頁起始頁
    per_page = 50

    while True:
        # 添加分頁參數
        params = {
            'per_page': per_page,  # 每頁的數量，最大值為 100
            'page': page     # 當前頁數
        }
        response = requests.get(url, headers=headers, params=params)
        if response.status_code != 200:
            print(f"Error: {response.status_code}, {response.json()}")
            break

        # 獲取當前頁數據
        data = response.json()
    
        if not data:  # 如果當前頁無數據，結束循環
            break

        members.extend(data)  # 添加數據到總列表
        if len(data) < per_page:
            break        
        page += 1  # 下一頁
    return members

def get_display_name(row):
    url_name = f"https://api.github.com/users/{row['帳號名稱']}"
    response_name = requests.get(url_name, headers=headers).json()
    logging.info(f"打API帳號: {row['帳號名稱']:10s} ")
    return response_name['name'] if response_name['name'] else response_name['login']

allToken = []
with open(DIR + "\Backend\.env", encoding='utf8') as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith('#'):
            key, value = line.split('=', 1)
            key = key[4:-6]
            allToken.append({'組織': key.strip(), 'token': value.strip()})

df_token = pd.DataFrame(allToken)[1:]
logging.info(f"共有{len(df_token)}個組織")

# df_OrgOwner = pd.DataFrame(columns=['組織', '帳號名稱', '存取人員'])
df_OrgOwner = pd.DataFrame(columns=['組織', '帳號名稱'])
# df_OrgOwner = pd.read_excel(MEMBER_FILE)
logging.info(f"{'-'*10}組織成員{'-'*10}")

for orgName, token in df_token.values:
    print(orgName)
    logging.info(f"開始處理 {orgName:35s} ")
    all_members = fetch_all_members(orgName, token)
    headers = {
        'Authorization': f'token {token}'
    }

    member = [all['login'] for all in all_members]
    # print(member)
    newrow = pd.DataFrame({'組織':[orgName]*len(member),
                           '帳號名稱':member,})
    df_OrgOwner = pd.concat([df_OrgOwner, newrow], ignore_index=True)
    # for i in range(len(member)):
    #     try:
    #         url_role = f"https://api.github.com/orgs/{orgName}/memberships/{member[i]}" # role
    #         response_role = requests.get(url_role, headers=headers).json()
    
    #         # get display name
    #         # url_name = f"https://api.github.com/users/{member[i]}" # name
    #         # response_name = requests.get(url_name, headers=headers).json()
            
    #         newrow = pd.DataFrame({'組織':[orgName], 
    #                             '帳號名稱':[member[i]],
    #                             # '存取人員':[response_name['name'] if response_name['name'] else response_name['login']], 
    #                             })
    #         print(i, newrow.values)
    #         df_OrgOwner = pd.concat([df_OrgOwner, newrow], ignore_index=True)

    #     except Exception as e:
    #         print(f'ERROR: \n {e}')
    #         df_OrgOwner.to_excel(MEMBER_FILE, index=False)
    #         logging.error(f"ERROR: {e}")
    #         logging.error(f'組織: {orgName}, 帳號: {member[i]} 處理錯誤')
    #         logging.error(response_role)
    #         break
    logging.info(f"組織 {orgName:35s} 共 {len(member)} 位成員")
    time.sleep(2)

df_OrgOwner.to_excel(MEMBER_FILE, index=False)
logging.info(f"組織成員獲取結束")


# outside collaborators
logging.info(f"{'-'*10}外部協作者{'-'*10}")
print('外部協作者')
for orgName, token in df_token.values:
    # print(orgName)
    logging.info(f"開始處理 {orgName:35s} ")
    headers = {
        'Authorization': f'token {token}'
    }
    url = f"https://api.github.com/orgs/{orgName}/outside_collaborators" 
    response_outside = requests.get(url, headers=headers).json()
    
    outside = [all['login'] for all in response_outside]
    newrow = pd.DataFrame({'組織':[orgName]*len(outside),
                           '帳號名稱':outside,})
    df_OrgOwner = pd.concat([df_OrgOwner, newrow], ignore_index=True)
    # if response_outside:
    #     print(orgName)    
           
    #     for i in response_outside:
    #         # url_name = f"https://api.github.com/users/{i['login']}" # name
    #         # response_name = requests.get(url_name, headers=headers).json()

    #         # get display name
    #         newrow = pd.DataFrame({'組織':[orgName], 
    #                             '帳號名稱' :[i['login']],
    #                             # '存取人員' : [response_name['name'] if response_name['name'] else response_name['login']],
    #                             })
    #         print(newrow.values)
    #         df_OrgOwner = pd.concat([df_OrgOwner, newrow], ignore_index=True)
    time.sleep(2)
    logging.info(f"組織 {orgName:35s} 共 {len(response_outside)} 位外部協作者")
df_OrgOwner.to_excel(MEMBER_FILE, index=False)
print('API request finished')

# mapping display name
df_before.drop(columns=['組織', '部門'], inplace=True)
df_before = df_before.drop_duplicates(subset=['帳號名稱'])

df_name = pd.DataFrame(columns=['帳號名稱'])
df_name['帳號名稱'] = df_OrgOwner['帳號名稱'].unique()

df_name = pd.merge(df_name, df_before, on=['帳號名稱'], how='left')
df_name['存取人員'] = df_name.apply( lambda row:
    get_display_name(row) if pd.isna(row['存取人員']) else row['存取人員'],
    axis=1
)
df_OrgOwner = pd.merge(df_OrgOwner, df_name, on=['帳號名稱'], how='left')

df_OrgOwner['存取人員'] = df_OrgOwner['存取人員'].apply(lambda x: ''.join(x.split(' ')) if isinstance(x, str) else x)
logging.info(f"存取人員名稱處理結束")
df_OrgOwner['組織'] = df_OrgOwner['組織'].apply(lambda x: df_dept.loc[df_dept['org_EN']==x, 'org_CN'].values[0] if x in df_dept['org_EN'].values else x)
logging.info(f"組織轉中文結束")
df_OrgOwner.to_excel(MEMBER_FILE, index=False)


# get department by display name
def getDept(name):
    # 用regex抓英文
    # print(name)
    name = name.strip()
    dept = re.findall(r'^[a-zA-Z0-9_-]+', name)
    if dept:
        en = dept[0].upper().replace('-', '_')
    else:
        en = name
    en = re.sub(r'_$', '', en)
    en = en.replace('ACD_RA2', 'ACD_RD2')
    en = en.replace('OLR_ART', 'OLD_ART')
    en = en.replace('ADM_MISSW', 'ADM_MIS')
    return en

def getDeptByOrg(df):
    df['部門'] = df['部門'].str.replace(r'_$', '', regex=True)
    df['部門'] = df.apply(
        lambda row: df_dept.loc[df_dept['org_CN'] == row['組織'], 'dep'].values[0]
        if row['部門'] not in df_dept['dep'].values #and row['組織'] in df_dept['org_CN'].values
        else row['部門']
        , axis=1
    )
    return df

df_OrgOwner['部門'] = df_OrgOwner['存取人員'].apply(getDept)
df_OrgOwner = getDeptByOrg(df_OrgOwner)
df_OrgOwner.to_excel(MEMBER_FILE, index=False)
logging.info(f"獲取存取人員部門")

# make table
df_OrgOwner = df_OrgOwner[['存取人員', '部門']].drop_duplicates()
# df_OrgOwner.to_excel('參考檔案\drop_dup.xlsx', index=False)

df_aggregate = df_OrgOwner.groupby('部門').agg({'存取人員': 'count'}).reset_index()
df_aggregate['存取人員百分比'] = df_aggregate['存取人員'].apply(lambda x: x / df_aggregate['存取人員'].sum())

df_sum = pd.DataFrame({'部門': ['總計'], 
                       '存取人員': df_aggregate['存取人員'].sum(), 
                       '存取人員百分比': df_aggregate['存取人員百分比'].sum()}
                       , index=[0])

df_aggregate = pd.concat([df_aggregate, df_sum], ignore_index=True)
logging.info(f"計算各部門人數比例")


# add % format
with pd.ExcelWriter(MEMBER_COUNT_FILE, engine='xlsxwriter') as writer:
    df_aggregate.to_excel(writer, index=False)
    
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']
    percent_format = workbook.add_format({'num_format': '0.00%'})
    worksheet.set_column('C:C', None, percent_format)

# edit excel style
workbook = load_workbook(MEMBER_FILE)
worksheet = workbook.active
for cell in worksheet[1]:
    cell.font = Font(bold=False) 
    cell.border = None
    cell.alignment = Alignment(horizontal='left')

    for col in worksheet.columns:
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = 16

workbook.save(MEMBER_FILE)
workbook.close()

workbook = load_workbook(MEMBER_COUNT_FILE)
worksheet = workbook.active
for cell in worksheet[1]:
    cell.font = Font(bold=False) 
    cell.border = None
    cell.alignment = Alignment(horizontal='left')

    for col in worksheet.columns:
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = 16

workbook.save(MEMBER_COUNT_FILE)
workbook.close()

def del_files(folder):
    for file in os.listdir(folder):
        file_path = os.path.join(folder, file)

        if os.path.isfile(file_path):
            file_mtime = os.path.getmtime(file_path)
            if file_mtime < N_DAYS_AGO:
                logging.info(f"刪除 {file_path}")
                os.remove(file_path)
        
        if os.path.isdir(file_path):
            del_files(file_path)
logging.info(f"刪除{N}天前編輯的檔案")
del_files(MEMBER_FOLDER)

# # upload to onedrive
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# import autoit

# # ChromeDriver 
# DIR = os.getcwd()
# CURRENTDAY = datetime.now().strftime("%Y%m%d")
# ACCOUNT = "dorispeiyu@gmail.com"
# PASSWORD = os.environ.get('ONEDRIVE_PASSWORD')
# # PASSWORD = 'Peiyu@610352'



# max_retries = 3
# retry_count = 0

# driver = webdriver.Chrome()
# driver.get("https://onedrive.live.com/login/")

# iframe = WebDriverWait(driver, 15).until(
#     EC.presence_of_element_located((By.TAG_NAME, "iframe"))
# )
# driver.switch_to.frame(iframe)

# # account
# email_input = WebDriverWait(driver, 15).until(
#     EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/main/div[2]/div[2]/div/input"))
# )
# time.sleep(3)
# email_input.send_keys(ACCOUNT)
# next_button = driver.find_element(By.XPATH, "/html/body/div[2]/div/main/div[2]/div[4]/input")
# next_button.click()
# logging.info(f"成功輸入帳號 : {ACCOUNT}")


# # pwd
# time.sleep(5)
# driver.switch_to.active_element.send_keys(PASSWORD)
# login_button = driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div/div[2]/div[1]/div/div/div/div/div[2]/div[2]/div/form/div[5]/div/div/div/div/button")
# login_button.click()
# logging.info("成功輸入密碼")

# # keep login window
# no_button = WebDriverWait(driver, 15).until(
#     EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div/div/div[2]/div[1]/div/div/div/div/div[2]/div[2]/div/form/div[3]/div[2]/div/div[1]/button"))
# )
# no_button.click()
# logging.info("登入成功")


# # in onedrive
# try: 
#     time.sleep(5)
#     share_link = WebDriverWait(driver, 15).until(
#         EC.presence_of_element_located((By.XPATH, "//a[@title='我的檔案']"))
#     )
#     share_link.click()
#     logging.info("成功進入我的檔案資料夾")

#     time.sleep(3)
#     shared_folder = WebDriverWait(driver, 30).until(
#         EC.presence_of_element_located((By.XPATH, "//button[@title='文件']"))
#     )
#     shared_folder.click()
#     logging.info("成功進入文件資料夾")

#     time.sleep(3)
#     github_folder = WebDriverWait(driver, 30).until(
#         EC.presence_of_element_located((By.XPATH, "//button[@title='Github']"))
#     )
#     github_folder.click()
#     logging.info("成功進入Github資料夾")
#     time.sleep(3)
#     acc_folder = WebDriverWait(driver, 30).until(
#         EC.presence_of_element_located((By.XPATH, "//button[@title='帳號清查']"))
#     )
#     acc_folder.click()
#     logging.info("成功進入帳號清查資料夾")

# except Exception as e:
#     logging.error(f"進入資料夾失敗: {e}")
#     sys.exit(1)

# try: # upload file
#     time.sleep(3)
#     add_button = WebDriverWait(driver, 15).until(
#         EC.element_to_be_clickable((By.XPATH, '//button[@data-id="AddNew" and @title="新增新的"]'))

#     )
#     add_button.click()
    
#     time.sleep(3)
#     file_upload = WebDriverWait(driver, 15).until(
#         EC.element_to_be_clickable((By.XPATH, '//button[@title="檔案上傳" and @data-automationid="UploadFileCommand"]'))
#     )
#     file_upload.click()

#     logging.info("成功點選上傳檔案按鈕，彈出win檔案選擇視窗")
#     time.sleep(3)

# except Exception as e:
#     driver.save_screenshot(os.path.join(LOG_FOLDER, f'log/{CURRENTDAY}_selectFolderError.png'))
#     logging.error("擷取螢幕畫面")    
#     logging.error(f"點選上傳檔案失敗: {e}")  
#     sys.exit(1)

# try:
#     time.sleep(3)
#     try:
#         autoit.win_activate("開啟") # 轉成活動狀態
#         # autoit.win_wait_active("開啟") # 英文系統要改
#         if not autoit.win_wait_active("開啟", 10):
#             autoit.win_activate("開啟")
#         logging.info('喚醒開啟視窗')
#     except Exception as e:
#         driver.save_screenshot(os.path.join(LOG_FOLDER, f'log/{CURRENTDAY}_activateWindowError.png'))
#         logging.error("擷取螢幕畫面")        
#         logging.error(f'喚醒開啟視窗錯誤{e}')
    

#     autoit.control_send("開啟", "Edit1", MEMBER_COUNT_FILE)
#     logging.info('輸入檔案位置')
#     time.sleep(3)
#     file_entered = autoit.control_get_text("開啟", "Edit1")
#     autoit.control_click("開啟", "Button1")
#     logging.info(f"點選開啟，成功在win視窗選擇檔案 : {file_entered}")
#     logging.info('任務執行完成 - 上傳部門Github帳號用量')
#     time.sleep(5)
    
# except Exception as e:
#     driver.save_screenshot(os.path.join(LOG_FOLDER, f'log/{CURRENTDAY}_selectFileError.png'))
#     logging.error("擷取螢幕畫面")
#     logging.error(f"在win視窗選擇檔案失敗: {e}")  
#     sys.exit(1)